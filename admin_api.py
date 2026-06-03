"""管理后台API"""
import io
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Blueprint, request, jsonify, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from models import db, User, AttendanceRecord, Holiday
from utils import (
    get_attendance_cycle, get_cycle_by_label, get_natural_month_cycle,
    get_day_type, calculate_work_and_overtime, calculate_overtime_pay,
    get_hourly_rate, get_default_record_for_date, count_workdays_in_range,
    today_bj, TZ
)

admin_bp = Blueprint("admin", __name__)


def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return jsonify({"code": 401, "msg": "请先登录"}), 401
        return f(*args, **kwargs)
    return decorated


# ==================== 登录 ====================

@admin_bp.route("/api/login", methods=["POST"])
def login():
    from config import Config
    data = request.get_json()
    password = data.get("password", "")
    if password == Config.ADMIN_PASSWORD:
        session["admin_logged_in"] = True
        return jsonify({"code": 0, "msg": "登录成功"})
    return jsonify({"code": 1, "msg": "密码错误"})


@admin_bp.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"code": 0, "msg": "已退出"})


@admin_bp.route("/api/check_login", methods=["GET"])
def check_login():
    return jsonify({"code": 0, "logged_in": session.get("admin_logged_in", False)})


# ==================== 用户管理 ====================

@admin_bp.route("/api/users", methods=["GET"])
@login_required
def get_users():
    users = User.query.order_by(User.id).all()
    return jsonify({
        "code": 0,
        "data": [u.to_dict() for u in users]
    })


@admin_bp.route("/api/users/<int:user_id>/salary", methods=["PUT"])
@login_required
def update_salary(user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    user.base_salary = float(data.get("base_salary", user.base_salary))
    db.session.commit()
    return jsonify({"code": 0, "msg": "底薪已更新"})


@admin_bp.route("/api/users/<int:user_id>/nickname", methods=["PUT"])
@login_required
def update_nickname(user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    user.nickname = data.get("nickname", "").strip()
    db.session.commit()
    return jsonify({"code": 0, "msg": "昵称已更新"})


# ==================== 考勤记录 ====================

@admin_bp.route("/api/records", methods=["GET"])
@login_required
def get_records():
    """获取考勤记录，支持按用户、时间范围筛选"""
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = AttendanceRecord.query

    if user_id:
        query = query.filter(AttendanceRecord.user_id == user_id)
    if start_date:
        query = query.filter(AttendanceRecord.date >= date.fromisoformat(start_date))
    if end_date:
        query = query.filter(AttendanceRecord.date <= date.fromisoformat(end_date))

    records = query.order_by(AttendanceRecord.date.desc(), AttendanceRecord.id).all()

    result = []
    for r in records:
        item = r.to_dict()
        user = User.query.get(r.user_id)
        item["nickname"] = user.nickname or f"用户{user.id}" if user else "未知"
        item["base_salary"] = user.base_salary if user else 0
        result.append(item)

    return jsonify({"code": 0, "data": result})


@admin_bp.route("/api/records/manual", methods=["POST"])
@login_required
def manual_record():
    """补卡"""
    data = request.get_json()
    user_id = data.get("user_id")
    record_date = date.fromisoformat(data.get("date"))
    check_in_str = data.get("check_in")  # HH:MM
    check_out_str = data.get("check_out")  # HH:MM
    remark = data.get("remark", "管理员补卡")

    user = User.query.get_or_404(user_id)

    check_in = datetime.combine(record_date, datetime.strptime(check_in_str, "%H:%M").time(), tzinfo=TZ) if check_in_str else None
    check_out = datetime.combine(record_date, datetime.strptime(check_out_str, "%H:%M").time(), tzinfo=TZ) if check_out_str else None

    # 查找或创建记录
    record = AttendanceRecord.query.filter_by(user_id=user_id, date=record_date).first()
    if not record:
        record = AttendanceRecord(user_id=user_id, date=record_date)
        db.session.add(record)

    record.check_in = check_in
    record.check_out = check_out
    record.is_manual = True
    record.remark = remark

    # 重新计算工时
    if check_in and check_out:
        day_type = get_day_type(record_date)
        work_hours, overtime_hours = calculate_work_and_overtime(check_in, check_out, day_type)
        hourly_rate = get_hourly_rate(user.base_salary)
        overtime_pay = calculate_overtime_pay(overtime_hours, day_type, hourly_rate)
        record.work_hours = work_hours
        record.overtime_hours = overtime_hours
        record.overtime_pay = overtime_pay

    db.session.commit()
    return jsonify({"code": 0, "msg": "补卡成功"})


@admin_bp.route("/api/records/<int:record_id>", methods=["DELETE"])
@login_required
def delete_record(record_id):
    record = AttendanceRecord.query.get_or_404(record_id)
    db.session.delete(record)
    db.session.commit()
    return jsonify({"code": 0, "msg": "已删除"})


# ==================== 考勤汇总 ====================

@admin_bp.route("/api/summary", methods=["GET"])
@login_required
def get_summary():
    """获取考勤汇总（带默认值填充）"""
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    cycle = request.args.get("cycle")  # 如 "2026年4月"
    mode = request.args.get("mode", "cycle")  # "cycle" | "natural"

    if cycle:
        start_date, end_date = get_cycle_by_label(cycle)
    elif start_date and end_date:
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)
    elif mode == "natural":
        start_date, end_date, _ = get_natural_month_cycle()
    else:
        start_date, end_date, _ = get_attendance_cycle()

    query = User.query
    if user_id:
        query = query.filter(User.id == user_id)
    users = query.all()

    result = []
    for user in users:
        records = AttendanceRecord.query.filter(
            AttendanceRecord.user_id == user.id,
            AttendanceRecord.date >= start_date,
            AttendanceRecord.date <= end_date
        ).all()

        record_map = {r.date: r for r in records}
        total_work = 0.0
        total_overtime = 0.0
        total_pay = 0.0
        details = []

        current = start_date
        while current <= end_date:
            if current in record_map:
                r = record_map[current]
                total_work += r.work_hours
                total_overtime += r.overtime_hours
                total_pay += r.overtime_pay
                details.append(r.to_dict())
            else:
                # 使用默认值
                default = get_default_record_for_date(current, user)
                total_work += default.work_hours
                total_overtime += default.overtime_hours
                total_pay += default.overtime_pay
                details.append(default.to_dict())
            current += timedelta(days=1)

        workday_count = count_workdays_in_range(start_date, end_date)

        result.append({
            "user_id": user.id,
            "nickname": user.nickname or f"用户{user.id}",
            "base_salary": user.base_salary,
            "workday_count": workday_count,
            "total_work_hours": round(total_work, 1),
            "total_overtime_hours": round(total_overtime, 1),
            "total_overtime_pay": round(total_pay, 2),
            "details": details,
        })

    return jsonify({
        "code": 0,
        "data": result,
        "period": {"start": start_date.strftime("%Y-%m-%d"), "end": end_date.strftime("%Y-%m-%d")}
    })


# ==================== 导出Excel ====================

@admin_bp.route("/api/export", methods=["GET"])
def export_excel():
    """导出考勤报表Excel"""
    user_id = request.args.get("user_id", type=int)
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    cycle = request.args.get("cycle")
    mode = request.args.get("mode", "cycle")  # "cycle" | "natural"

    if cycle:
        start_date, end_date = get_cycle_by_label(cycle)
    elif start_date and end_date:
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)
    elif mode == "natural":
        start_date, end_date, _ = get_natural_month_cycle()
    else:
        start_date, end_date, _ = get_attendance_cycle()

    query = User.query
    if user_id:
        query = query.filter(User.id == user_id)
    users = query.order_by(User.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤汇总"

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    money_fmt = '#,##0.00'

    # 汇总表头
    summary_headers = ["姓名", "底薪(元/月)", "出勤天数", "总工时(h)", "加班时长(h)", "加班费(元)"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row = 2
    for user in users:
        records = AttendanceRecord.query.filter(
            AttendanceRecord.user_id == user.id,
            AttendanceRecord.date >= start_date,
            AttendanceRecord.date <= end_date
        ).all()
        record_map = {r.date: r for r in records}

        total_work = 0.0
        total_overtime = 0.0
        total_pay = 0.0

        current = start_date
        while current <= end_date:
            if current in record_map:
                r = record_map[current]
                total_work += r.work_hours
                total_overtime += r.overtime_hours
                total_pay += r.overtime_pay
            else:
                default = get_default_record_for_date(current, user)
                total_work += default.work_hours
                total_overtime += default.overtime_hours
                total_pay += default.overtime_pay
            current += timedelta(days=1)

        workday_count = count_workdays_in_range(start_date, end_date)
        nickname = user.nickname or f"用户{user.id}"

        values = [nickname, user.base_salary, workday_count,
                  round(total_work, 1), round(total_overtime, 1), round(total_pay, 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 6:  # 加班费列
                cell.number_format = money_fmt
        row += 1

    # 调整列宽
    col_widths = [12, 15, 10, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # ---- 明细表 ----
    ws2 = wb.create_sheet("考勤明细")

    detail_headers = ["姓名", "日期", "星期", "类型", "上班时间", "下班时间", "工时(h)", "加班(h)", "加班费(元)", "备注"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    week_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    type_names = {"workday": "工作日", "weekend": "休息日", "holiday": "法定节假日"}

    row = 2
    for user in users:
        records = AttendanceRecord.query.filter(
            AttendanceRecord.user_id == user.id,
            AttendanceRecord.date >= start_date,
            AttendanceRecord.date <= end_date
        ).order_by(AttendanceRecord.date).all()
        record_map = {r.date: r for r in records}
        nickname = user.nickname or f"用户{user.id}"

        current = start_date
        while current <= end_date:
            if current in record_map:
                r = record_map[current]
                ci = r.check_in.strftime('%H:%M') if r.check_in else "-"
                co = r.check_out.strftime('%H:%M') if r.check_out else "-"
                wh = round(r.work_hours, 1)
                oh = round(r.overtime_hours, 1)
                op = round(r.overtime_pay, 2)
                remark = r.remark or ""
            else:
                default = get_default_record_for_date(current, user)
                ci = default.check_in.strftime('%H:%M') if default.check_in else "-"
                co = default.check_out.strftime('%H:%M') if default.check_out else "-"
                wh = round(default.work_hours, 1)
                oh = round(default.overtime_hours, 1)
                op = round(default.overtime_pay, 2)
                remark = default.remark or ""

            day_type = get_day_type(current)

            values = [
                nickname,
                current.strftime('%Y-%m-%d'),
                week_names[current.weekday()],
                type_names.get(day_type, ""),
                ci, co, wh, oh, op, remark
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if col == 9:
                    cell.number_format = money_fmt
            row += 1

    # 调整列宽
    detail_widths = [12, 14, 8, 12, 10, 10, 10, 10, 14, 20]
    for i, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"考勤报表_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ==================== 节假日管理 ====================

@admin_bp.route("/api/holidays", methods=["GET"])
@login_required
def get_holidays():
    year = request.args.get("year", type=int, default=today_bj().year)
    holidays = Holiday.query.filter_by(year=year).order_by(Holiday.date).all()
    return jsonify({"code": 0, "data": [h.to_dict() for h in holidays]})


@admin_bp.route("/api/holidays", methods=["POST"])
@login_required
def add_holiday():
    data = request.get_json()
    holiday_date = date.fromisoformat(data.get("date"))
    name = data.get("name", "")

    existing = Holiday.query.filter_by(date=holiday_date).first()
    if existing:
        return jsonify({"code": 1, "msg": "该日期已存在"})

    holiday = Holiday(date=holiday_date, name=name, year=holiday_date.year)
    db.session.add(holiday)
    db.session.commit()
    return jsonify({"code": 0, "msg": "添加成功"})


@admin_bp.route("/api/holidays/<int:holiday_id>", methods=["DELETE"])
@login_required
def delete_holiday(holiday_id):
    holiday = Holiday.query.get_or_404(holiday_id)
    db.session.delete(holiday)
    db.session.commit()
    return jsonify({"code": 0, "msg": "已删除"})


@admin_bp.route("/api/holidays/batch", methods=["POST"])
@login_required
def batch_add_holidays():
    """批量添加节假日"""
    data = request.get_json()
    dates = data.get("dates", [])  # ["2026-01-01", "2026-01-02", ...]
    name = data.get("name", "")

    count = 0
    for d_str in dates:
        d = date.fromisoformat(d_str)
        if not Holiday.query.filter_by(date=d).first():
            holiday = Holiday(date=d, name=name, year=d.year)
            db.session.add(holiday)
            count += 1

    db.session.commit()
    return jsonify({"code": 0, "msg": f"成功添加 {count} 个节假日"})


# ==================== 考勤周期信息 ====================

@admin_bp.route("/api/cycle", methods=["GET"])
@login_required
def get_cycle_info():
    start, end, label = get_attendance_cycle()
    return jsonify({
        "code": 0,
        "data": {
            "current": {"start": start.strftime("%Y-%m-%d"), "end": end.strftime("%Y-%m-%d"), "label": label}
        }
    })
